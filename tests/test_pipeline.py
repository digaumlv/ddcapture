"""Ponta a ponta: dashboard -> valores -> arquivos, com cliente falso."""

from __future__ import annotations

import json
import sqlite3

import pytest
from openpyxl import load_workbook

from ddcapture.runner import capturar, preparar
from ddcapture.sinks import gravar_todos

INICIO, FIM = 1_700_000_000, 1_700_003_600


class ClienteFalso:
    """Responde a todos os endpoints que a fixture exercita."""

    def __init__(self):
        self.chamadas: list[str] = []

    def get(self, caminho, params=None):
        self.chamadas.append(caminho)
        if caminho == "/api/v1/query":
            return {
                "series": [
                    {
                        "metric": "aws.rds.database_connections",
                        "scope": "env:prod",
                        "pointlist": [[1_700_000_000_000, 42.0]],
                        "unit": [{"short_name": "conn"}],
                    }
                ]
            }
        if caminho.startswith("/api/v1/slo/"):
            return {
                "data": {
                    "slo": {"name": "Checkout disponivel"},
                    "overall": {"sli_value": 99.87, "target": 99.9},
                }
            }
        if caminho == "/api/v1/monitor":
            return [{"overall_state": "OK"}, {"overall_state": "Alert"}]
        raise AssertionError(f"GET inesperado: {caminho}")

    def post(self, caminho, corpo):
        self.chamadas.append(caminho)
        if caminho == "/api/v2/query/scalar":
            return {
                "data": {
                    "attributes": {
                        "columns": [
                            {"name": "v", "type": "number", "values": [7.5],
                             "meta": {"unit": [{"short_name": "un"}]}}
                        ]
                    }
                }
            }
        if caminho == "/api/v2/query/timeseries":
            return {
                "data": {
                    "attributes": {
                        "series": [{"group_tags": ["host:web-01"], "query_index": 0}],
                        "times": [1_700_000_000_000],
                        "values": [[3.25]],
                    }
                }
            }
        raise AssertionError(f"POST inesperado: {caminho}")


@pytest.fixture
def resultado(dashboard, config):
    r = preparar(dashboard, config, INICIO, FIM)
    return capturar(ClienteFalso(), r, config)


def test_dry_run_nao_toca_a_rede(dashboard, config):
    """preparar() e o que o --dry-run executa: inventario sem consultar dados."""
    r = preparar(dashboard, config, INICIO, FIM)

    assert r.dashboard_id == "abc-def-ghi"
    assert len(r.widgets) == 10  # 9 com dado + a nota
    assert sum(1 for w in r.widgets if w.sem_query) == 1
    assert len(r.specs) == 9
    # Nenhuma medicao ainda.
    assert r.medicoes == []


def test_todos_os_widgets_com_dado_produzem_valor(resultado):
    assert resultado.capturados, "nenhum valor capturado"
    assert not resultado.falhas, [m.erro for m in resultado.falhas]


def test_falha_nao_poe_a_query_na_coluna_campo(dashboard, config):
    """O campo mantem o titulo do widget mesmo quando a leitura falha.

    Antes, o nome caia para o texto da query e a coluna Campo virava
    'Total de Reenvios | (query)'.
    """

    class ClienteQueFalha(ClienteFalso):
        def post(self, caminho, corpo):
            if caminho == "/api/v2/query/scalar":
                from ddcapture.client import ErroApi

                raise ErroApi("403 - sem permissao", 403)
            return super().post(caminho, corpo)

    r = capturar(ClienteQueFalha(), preparar(dashboard, config, INICIO, FIM), config)
    falha = next(m for m in r.falhas if m.widget_titulo == "Memoria livre")

    assert falha.nome_valor == "Memoria livre"
    assert "(query)" not in falha.nome_valor
    assert "|" not in falha.nome_valor


def test_cada_valor_tem_nome_e_categoria(resultado):
    for m in resultado.medicoes:
        assert m.nome_valor.strip(), f"valor sem nome no widget {m.widget_titulo}"
        assert m.categoria.strip()
        assert m.categoria_origem in {"grupo", "palavra-chave", "namespace", "fallback"}


def test_nome_combina_titulo_do_widget_com_a_serie(resultado):
    cpu = next(m for m in resultado.medicoes if m.widget_titulo == "CPU por host")
    # O alias da formula ('CPU %') e um rotulo humano: entra no nome.
    assert cpu.nome_valor.startswith("CPU por host | CPU %")
    # O escopo da serie entra no nome.
    assert "host:web-01" in cpu.nome_valor


def test_formula_sem_alias_nao_polui_o_nome(dashboard, config):
    """Formulas sem alias - comum em dashboards criados pela UI.

    A API devolve a expressao ('default_zero(query1)') como nome da coluna.
    Isso nao identifica nada - o rotulo util e o titulo do widget.
    """

    class ClienteFormulaCrua(ClienteFalso):
        def post(self, caminho, corpo):
            if caminho == "/api/v2/query/scalar":
                return {
                    "data": {
                        "attributes": {
                            "columns": [
                                {
                                    "name": "default_zero(query1)",
                                    "type": "number",
                                    "values": [42.0],
                                }
                            ]
                        }
                    }
                }
            return super().post(caminho, corpo)

    # Uma formula sem alias, como a UI do Datadog costuma gerar.
    for w in dashboard["widgets"][0]["definition"]["widgets"]:
        if w["definition"].get("title") == "Memoria livre":
            req = w["definition"]["requests"][0]
            req["formulas"] = [{"formula": "default_zero(query1)"}]

    r = capturar(ClienteFormulaCrua(), preparar(dashboard, config, INICIO, FIM), config)
    mem = next(m for m in r.medicoes if m.widget_titulo == "Memoria livre")

    assert mem.valor == 42.0
    assert mem.nome_valor == "Memoria livre"
    assert "default_zero" not in mem.nome_valor


def test_as_quatro_camadas_aparecem_no_resultado(resultado):
    origens = {m.categoria_origem for m in resultado.medicoes}
    assert origens == {"grupo", "palavra-chave", "namespace", "fallback"}


def test_agrupamento_por_categoria(resultado):
    categorias = resultado.por_categoria()
    assert "Infraestrutura AWS" in categorias
    assert "RDS" in categorias
    assert "Erros" in categorias
    assert "APM" in categorias
    assert "sem-categoria" in categorias


def test_tags_viram_dimensao(resultado):
    mem = next(m for m in resultado.medicoes if m.widget_titulo == "Memoria livre")
    assert mem.tags.get("env") == "prod"


def test_falha_de_uma_query_nao_derruba_a_execucao(dashboard, config):
    class ClienteQueFalha(ClienteFalso):
        def post(self, caminho, corpo):
            if caminho == "/api/v2/query/scalar":
                from ddcapture.client import ErroApi

                raise ErroApi("403 - sem permissao", 403)
            return super().post(caminho, corpo)

    r = capturar(ClienteQueFalha(), preparar(dashboard, config, INICIO, FIM), config)

    assert r.falhas, "esperava falhas registradas"
    assert r.capturados, "as demais queries deveriam continuar"
    # A falha fica no registro, com valor nulo.
    assert all(m.valor is None for m in r.falhas)


# --- Saidas -------------------------------------------------------------


@pytest.fixture
def escritos(resultado, config, tmp_path):
    config.saida_dir = tmp_path
    # Liga todos os sinks explicitamente: o settings.yaml deixa so o xlsx
    # ativo, mas os demais continuam suportados e precisam de cobertura.
    config.sinks = {
        "json": True,
        "csv": True,
        "xlsx": True,
        "xlsx_por_widget": True,
        "sqlite": True,
    }
    return gravar_todos(resultado, config)


def test_grava_todos_os_formatos(escritos):
    arquivos = [c for c in escritos if c.is_file()]
    pastas = [c for c in escritos if c.is_dir()]

    assert {c.suffix for c in arquivos} == {".json", ".csv", ".xlsx", ".sqlite"}
    assert all(c.stat().st_size > 0 for c in arquivos)

    # O sink por widget devolve a pasta, nao um arquivo.
    assert len(pastas) == 1
    assert pastas[0].name.endswith("_widgets")
    assert list(pastas[0].glob("*.xlsx"))


def test_json_preserva_a_hierarquia(escritos, resultado):
    caminho = next(c for c in escritos if c.suffix == ".json")
    dados = json.loads(caminho.read_text(encoding="utf-8"))

    assert dados["dashboard"]["id"] == "abc-def-ghi"
    assert dados["resumo"]["valores_capturados"] == len(resultado.capturados)
    # O inventario inclui a nota, que nao gera valor.
    assert len(dados["inventario_widgets"]) == len(resultado.widgets)
    assert any(w["sem_query"] for w in dados["inventario_widgets"])

    categorias = {c["categoria"] for c in dados["categorias"]}
    assert "Infraestrutura AWS" in categorias
    # Dentro da categoria, os valores ficam sob o widget que os produziu.
    infra = next(c for c in dados["categorias"] if c["categoria"] == "Infraestrutura AWS")
    assert infra["widgets"][0]["valores"]


def test_csv_tem_uma_linha_por_valor(escritos, resultado):
    caminho = next(c for c in escritos if c.suffix == ".csv")
    linhas = caminho.read_text(encoding="utf-8-sig").strip().splitlines()

    assert len(linhas) == len(resultado.medicoes) + 1  # + cabecalho
    cabecalho = linhas[0].split(";")
    assert "nome_valor" in cabecalho and "categoria" in cabecalho
    # As tags de dimensao viram colunas proprias.
    assert {"env", "service", "team"} <= set(cabecalho)


def test_xlsx_tem_exatamente_uma_aba_por_categoria(escritos, resultado):
    caminho = next(c for c in escritos if c.suffix == ".xlsx")
    wb = load_workbook(caminho)

    categorias = resultado.por_categoria()
    # Nada de abas de servico - so as categorias.
    assert len(wb.sheetnames) == len(categorias)
    assert "RDS" in wb.sheetnames
    assert "Infraestrutura AWS" in wb.sheetnames


def test_xlsx_traz_so_campo_e_valor(escritos, resultado):
    """Sem dados adicionais: nem query, nem fonte, nem tags, nem timestamp."""
    caminho = next(c for c in escritos if c.suffix == ".xlsx")
    ws = load_workbook(caminho)["Infraestrutura AWS"]

    assert [c.value for c in ws[1]] == ["Campo", "Valor"]
    assert ws.max_column == 2

    esperadas = len(resultado.por_categoria()["Infraestrutura AWS"])
    assert ws.max_row == esperadas + 1  # + cabecalho


def test_xlsx_grava_valor_como_numero(escritos):
    """Texto nao soma no Excel."""
    caminho = next(c for c in escritos if c.suffix == ".xlsx")
    ws = load_workbook(caminho)["Infraestrutura AWS"]

    campo, valor = ws.cell(row=2, column=1).value, ws.cell(row=2, column=2).value
    assert isinstance(campo, str) and campo.strip()
    assert isinstance(valor, (int, float))


class ClienteSemDados(ClienteFalso):
    """Responde com sucesso, mas sem nenhuma serie/coluna na janela."""

    def post(self, caminho, corpo):
        if caminho == "/api/v2/query/scalar":
            return {"data": {"attributes": {"columns": []}}}
        if caminho == "/api/v2/query/timeseries":
            return {"data": {"attributes": {"series": [], "times": [], "values": []}}}
        return super().post(caminho, corpo)


class ClienteQueFalha(ClienteFalso):
    """A API recusa a query - nao sabemos quanto vale o campo."""

    def post(self, caminho, corpo):
        if caminho == "/api/v2/query/scalar":
            from ddcapture.client import ErroApi

            raise ErroApi("400 - Invalid query input", 400)
        return super().post(caminho, corpo)


def test_consulta_vazia_vira_zero(dashboard, config):
    """Query rodou e nao achou nada = zero ocorrencias."""
    r = capturar(ClienteSemDados(), preparar(dashboard, config, INICIO, FIM), config)

    assert r.sem_dados, "esperava campos marcados como sem dados"
    assert all(m.valor == 0.0 for m in r.sem_dados)
    # Sem dados nao e falha: a consulta funcionou.
    assert all(m.erro is None for m in r.sem_dados)


def test_zero_por_ausencia_e_distinguivel_de_zero_medido(dashboard, config):
    r_vazio = capturar(ClienteSemDados(), preparar(dashboard, config, INICIO, FIM), config)
    r_normal = capturar(ClienteFalso(), preparar(dashboard, config, INICIO, FIM), config)

    mem_vazio = next(m for m in r_vazio.medicoes if m.widget_titulo == "Memoria livre")
    mem_normal = next(m for m in r_normal.medicoes if m.widget_titulo == "Memoria livre")

    assert mem_vazio.valor == 0.0 and mem_vazio.sem_dados is True
    # O valor 7.5 veio da API: nao e sem_dados, mesmo que fosse 0.
    assert mem_normal.valor == 7.5 and mem_normal.sem_dados is False


def test_erro_de_api_nao_vira_zero(dashboard, config):
    """Zero aqui seria mentira: a query foi recusada, o valor e desconhecido."""
    r = capturar(ClienteQueFalha(), preparar(dashboard, config, INICIO, FIM), config)

    assert r.falhas
    assert all(m.valor is None for m in r.falhas)
    assert all(m.erro for m in r.falhas)
    # Falha nao entra na contagem de sem_dados.
    assert not any(m in r.sem_dados for m in r.falhas)


def test_xlsx_traz_zero_para_consulta_vazia(dashboard, config, tmp_path):
    config.saida_dir = tmp_path
    config.sinks = {"xlsx": True}
    r = capturar(ClienteSemDados(), preparar(dashboard, config, INICIO, FIM), config)
    caminho = gravar_todos(r, config)[0]

    ws = load_workbook(caminho)["Infraestrutura AWS"]
    valores = [ws.cell(row=l, column=2).value for l in range(2, ws.max_row + 1)]

    assert 0 in valores or 0.0 in valores
    assert None not in valores, "consulta vazia deveria ter virado 0"


def test_xlsx_deixa_celula_vazia_quando_a_api_recusa(dashboard, config, tmp_path):
    """Campo existe mas nao pode ser lido: vazio, nunca zero."""
    config.saida_dir = tmp_path
    config.sinks = {"xlsx": True}
    r = capturar(ClienteQueFalha(), preparar(dashboard, config, INICIO, FIM), config)
    caminho = gravar_todos(r, config)[0]

    wb = load_workbook(caminho)
    vazios = [
        wb[nome].cell(row=linha, column=2).value
        for nome in wb.sheetnames
        for linha in range(2, wb[nome].max_row + 1)
        if wb[nome].cell(row=linha, column=2).value in (None, "")
    ]
    assert vazios, "as falhas deveriam aparecer com celula vazia"


def test_sqlite_acumula_snapshots(escritos, resultado, config):
    caminho = next(c for c in escritos if c.suffix == ".sqlite")

    # Segunda execucao no mesmo arquivo.
    gravar_todos(resultado, config)

    conexao = sqlite3.connect(caminho)
    try:
        execucoes = conexao.execute("SELECT COUNT(*) FROM execucoes").fetchone()[0]
        medicoes = conexao.execute("SELECT COUNT(*) FROM medicoes").fetchone()[0]
        historico = conexao.execute("SELECT COUNT(*) FROM historico").fetchone()[0]
    finally:
        conexao.close()

    # Duas execucoes, nao uma sobrescrita.
    assert execucoes == 2
    assert medicoes == len(resultado.medicoes) * 2
    assert historico == medicoes
