"""Saida de um arquivo XLSX por widget."""

from __future__ import annotations

import pytest
from openpyxl import load_workbook

from ddcapture.sinks import excel_widget_sink
from test_pipeline import FIM, INICIO, ClienteFalso  # noqa: F401

from ddcapture.runner import capturar, preparar


@pytest.fixture
def resultado(dashboard, config):
    return capturar(ClienteFalso(), preparar(dashboard, config, INICIO, FIM), config)


@pytest.fixture
def pasta(resultado, config, tmp_path):
    config.saida_dir = tmp_path
    return excel_widget_sink.gravar(resultado, config)


def test_um_arquivo_por_widget_mais_o_indice(pasta, resultado):
    arquivos = sorted(p.name for p in pasta.glob("*.xlsx"))
    widgets_com_valor = {m.widget_id for m in resultado.medicoes}

    assert "_indice.xlsx" in arquivos
    assert len(arquivos) == len(widgets_com_valor) + 1


def test_nome_do_arquivo_carrega_titulo_e_id(pasta):
    nomes = [p.name for p in pasta.glob("*.xlsx")]
    cpu = next(n for n in nomes if n.startswith("CPU por host"))
    # O widget_id entra sempre: titulos se repetem no mesmo dashboard.
    assert cpu.endswith(".xlsx")
    assert "__" in cpu


def test_titulos_repetidos_nao_se_sobrescrevem(dashboard, config, tmp_path):
    """E comum um dashboard repetir o mesmo titulo em varios widgets."""
    grupo = dashboard["widgets"][0]["definition"]["widgets"]
    # Dois widgets distintos com o mesmo titulo.
    grupo[0]["definition"]["title"] = "Duplicado"
    grupo[1]["definition"]["title"] = "Duplicado"

    config.saida_dir = tmp_path
    r = capturar(ClienteFalso(), preparar(dashboard, config, INICIO, FIM), config)
    destino = excel_widget_sink.gravar(r, config)

    duplicados = [p for p in destino.glob("Duplicado__*.xlsx")]
    assert len(duplicados) == 2, "um arquivo sobrescreveu o outro"


def test_arquivo_tem_metadados_e_tabela(pasta):
    caminho = next(p for p in pasta.glob("CPU por host*.xlsx"))
    ws = load_workbook(caminho)["Valores"]

    rotulos = [ws.cell(row=r, column=1).value for r in range(1, 11)]
    assert "Dashboard" in rotulos
    assert "Widget" in rotulos
    assert "Categoria" in rotulos
    assert "Query" in rotulos

    # Depois do bloco de metadados vem o cabecalho da tabela.
    linhas = list(ws.iter_rows(values_only=True))
    cabecalho = next(l for l in linhas if l and l[0] == "nome_valor")
    assert "valor" in cabecalho
    assert "unidade" in cabecalho
    # As tags de dimensao viram colunas.
    assert "env" in cabecalho


def test_valor_e_gravado_como_numero(pasta):
    caminho = next(p for p in pasta.glob("Memoria livre*.xlsx"))
    ws = load_workbook(caminho)["Valores"]

    linhas = list(ws.iter_rows(values_only=True))
    inicio = next(i for i, l in enumerate(linhas) if l and l[0] == "nome_valor")
    primeira = linhas[inicio + 1]

    # Coluna 'valor' - numero de verdade, nao texto, senao o Excel nao soma.
    assert isinstance(primeira[1], (int, float))
    assert primeira[1] == 7.5


def test_indice_lista_todos_os_arquivos(pasta):
    ws = load_workbook(pasta / "_indice.xlsx")["Indice"]

    linhas = list(ws.iter_rows(values_only=True))
    inicio = next(i for i, l in enumerate(linhas) if l and l[0] == "arquivo")
    dados = [l for l in linhas[inicio + 1 :] if l and l[0]]

    arquivos_na_pasta = {p.name for p in pasta.glob("*.xlsx")} - {"_indice.xlsx"}
    assert {str(l[0]) for l in dados} == arquivos_na_pasta

    cabecalho = linhas[inicio]
    assert "categoria" in cabecalho and "valores" in cabecalho and "query" in cabecalho


def test_nome_de_arquivo_sanitiza_caracteres_proibidos(dashboard, config, tmp_path):
    grupo = dashboard["widgets"][0]["definition"]["widgets"]
    grupo[0]["definition"]["title"] = 'Taxa: erros/total <critico>?'

    config.saida_dir = tmp_path
    r = capturar(ClienteFalso(), preparar(dashboard, config, INICIO, FIM), config)
    destino = excel_widget_sink.gravar(r, config)

    nomes = [p.name for p in destino.glob("*.xlsx")]
    alvo = next(n for n in nomes if n.startswith("Taxa"))
    assert not any(c in alvo for c in '<>:"/\\|?*')
    # O arquivo abre - prova que o nome e valido para o sistema de arquivos.
    load_workbook(destino / alvo)
