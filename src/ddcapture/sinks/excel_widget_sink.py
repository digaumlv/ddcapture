"""Saida XLSX granular: um arquivo por widget.

O excel_sink consolida tudo num arquivo com uma aba por categoria. Aqui e o
oposto: cada widget vira um arquivo proprio, para quem precisa circular o
recorte de um painel especifico sem mandar a planilha inteira.

Os arquivos vao para uma subpasta por execucao, com um _indice.xlsx que mapeia
widget -> arquivo. Sem o indice, um dashboard grande gera centenas de arquivos
com titulos que se repetem, e achar o certo vira garimpo.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..config import Config
from ..models import Measurement
from ..runner import Resultado
from . import prefixo_arquivo

_CABECALHO_FILL = PatternFill("solid", fgColor="1F2A44")
_CABECALHO_FONT = Font(color="FFFFFF", bold=True)
_ROTULO_FONT = Font(bold=True)
_ERRO_FILL = PatternFill("solid", fgColor="FDE7E9")

# Caracteres proibidos em nome de arquivo no Windows.
_INVALIDO_ARQUIVO = re.compile(r'[<>:"/\\|?*\x00-\x1f]')

# Colunas da tabela de valores. O que e constante no widget (titulo, grupo,
# query) fica no bloco de metadados, nao repetido linha a linha.
_COLUNAS = ["nome_valor", "valor", "unidade", "timestamp", "erro"]


def gravar(resultado: Resultado, config: Config) -> Path:
    """Escreve um arquivo por widget. Devolve a pasta criada."""
    pasta = config.saida_dir / f"{prefixo_arquivo(resultado)}_widgets"
    pasta.mkdir(parents=True, exist_ok=True)

    chaves_tag = config.tags_dimensao
    por_widget = _agrupar_por_widget(resultado)

    usados: set[str] = set()
    indice: list[dict[str, object]] = []

    for chave, medicoes in por_widget.items():
        primeira = medicoes[0]
        nome_arquivo = _nome_arquivo(primeira, chave, usados)
        caminho = pasta / nome_arquivo

        wb = Workbook()
        _escrever_widget(wb.active, primeira, medicoes, resultado, chaves_tag)
        wb.save(caminho)

        indice.append(
            {
                "arquivo": nome_arquivo,
                "widget": primeira.widget_titulo,
                "tipo": primeira.widget_tipo,
                "categoria": primeira.categoria,
                "grupo": primeira.grupo_pai or "",
                "valores": sum(1 for m in medicoes if m.erro is None),
                "falhas": sum(1 for m in medicoes if m.erro),
                "widget_id": primeira.widget_id,
                "query": primeira.query,
            }
        )

    _escrever_indice(pasta, indice, resultado)
    return pasta


def _agrupar_por_widget(resultado: Resultado) -> dict[str, list[Measurement]]:
    """Agrupa por widget_id - o titulo se repete no mesmo dashboard."""
    agrupado: dict[str, list[Measurement]] = {}
    for m in resultado.medicoes:
        agrupado.setdefault(m.widget_id or m.widget_titulo, []).append(m)
    return agrupado


def _escrever_widget(
    ws,
    primeira: Measurement,
    medicoes: list[Measurement],
    resultado: Resultado,
    chaves_tag: list[str],
) -> None:
    ws.title = "Valores"

    metadados = [
        ("Dashboard", f"{resultado.dashboard_titulo} ({resultado.dashboard_id})"),
        ("Widget", primeira.widget_titulo),
        ("Tipo", primeira.widget_tipo),
        ("Grupo", " > ".join(primeira.caminho_grupos) or "(sem grupo)"),
        ("Categoria", f"{primeira.categoria}  [{primeira.categoria_origem}]"),
        ("Fonte", primeira.data_source),
        ("Agregador", primeira.agregador or "-"),
        ("Query", primeira.query),
        ("Janela (epoch s)", f"{resultado.inicio_s} -> {resultado.fim_s}"),
        ("Capturado em", primeira.capturado_em),
    ]
    for rotulo, valor in metadados:
        ws.append([rotulo, _celula(valor)])
        ws.cell(row=ws.max_row, column=1).font = _ROTULO_FONT

    ws.append([])

    colunas = _COLUNAS + chaves_tag
    linha_cabecalho = ws.max_row + 1
    ws.append(colunas)
    for celula in ws[linha_cabecalho]:
        celula.font = _CABECALHO_FONT
        celula.fill = _CABECALHO_FILL
        celula.alignment = Alignment(vertical="center")

    linhas = [m.para_linha(chaves_tag) for m in medicoes]
    for linha in linhas:
        ws.append([_celula(linha.get(c)) for c in colunas])
        if linha.get("erro"):
            for celula in ws[ws.max_row]:
                celula.fill = _ERRO_FILL

    # Congela abaixo do cabecalho da tabela, nao no topo: o bloco de metadados
    # nao precisa ficar preso na tela.
    ws.freeze_panes = f"A{linha_cabecalho + 1}"
    ws.auto_filter.ref = f"A{linha_cabecalho}:{_coluna_letra(len(colunas))}{ws.max_row}"
    _ajustar_larguras(ws, colunas, linhas)
    # A coluna A tambem carrega os rotulos dos metadados.
    ws.column_dimensions["A"].width = max(
        ws.column_dimensions["A"].width or 10, 22
    )


def _escrever_indice(pasta: Path, indice: list[dict], resultado: Resultado) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Indice"

    ws.append(["Dashboard", f"{resultado.dashboard_titulo} ({resultado.dashboard_id})"])
    ws.append(["Arquivos", len(indice)])
    ws.append(["Janela (epoch s)", f"{resultado.inicio_s} -> {resultado.fim_s}"])
    for linha in range(1, 4):
        ws.cell(row=linha, column=1).font = _ROTULO_FONT
    ws.append([])

    colunas = [
        "arquivo", "widget", "tipo", "categoria", "grupo",
        "valores", "falhas", "widget_id", "query",
    ]
    linha_cabecalho = ws.max_row + 1
    ws.append(colunas)
    for celula in ws[linha_cabecalho]:
        celula.font = _CABECALHO_FONT
        celula.fill = _CABECALHO_FILL

    ordenado = sorted(indice, key=lambda d: (str(d["categoria"]), str(d["widget"])))
    for item in ordenado:
        ws.append([_celula(item.get(c)) for c in colunas])

    ws.freeze_panes = f"A{linha_cabecalho + 1}"
    ws.auto_filter.ref = f"A{linha_cabecalho}:{_coluna_letra(len(colunas))}{ws.max_row}"
    _ajustar_larguras(ws, colunas, ordenado)

    wb.save(pasta / "_indice.xlsx")


def _nome_arquivo(primeira: Measurement, chave: str, usados: set[str]) -> str:
    """Nome de arquivo unico e valido no Windows.

    O widget_id entra sempre: e comum um dashboard ter varios widgets com o
    mesmo titulo, e sem o id um sobrescreveria o outro em silencio.
    """
    base = _INVALIDO_ARQUIVO.sub("-", primeira.widget_titulo or "widget").strip(" .")
    base = re.sub(r"\s+", " ", base) or "widget"
    # Caminho no Windows estoura em 260 caracteres; a pasta ja consome boa parte.
    base = base[:60].strip(" .")

    sufixo = _INVALIDO_ARQUIVO.sub("-", str(chave))[:20]
    nome = f"{base}__{sufixo}.xlsx"

    # Cinto de seguranca: se ainda assim colidir, numera.
    contador = 2
    while nome.lower() in usados:
        nome = f"{base}__{sufixo}_{contador}.xlsx"
        contador += 1
    usados.add(nome.lower())
    return nome


def _coluna_letra(indice: int) -> str:
    return get_column_letter(indice)


def _celula(valor):
    """Converte para um tipo que o openpyxl aceita, mantendo numeros numericos."""
    if valor is None:
        return ""
    if isinstance(valor, (int, float, str)):
        return valor
    return str(valor)


def _ajustar_larguras(ws, colunas: list[str], linhas: list[dict]) -> None:
    for i, coluna in enumerate(colunas, start=1):
        maior = len(coluna)
        # Amostra as 200 primeiras linhas: varrer tudo nao muda o resultado.
        for linha in linhas[:200]:
            maior = max(maior, len(str(linha.get(coluna) or "")))
        ws.column_dimensions[get_column_letter(i)].width = min(max(maior + 2, 10), 60)
