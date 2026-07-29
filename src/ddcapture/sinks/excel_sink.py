"""Saida XLSX: um unico arquivo, uma aba por categoria, duas colunas.

Enxuto de proposito - so o nome do campo e o valor. Metadados (query, fonte,
tags, timestamp, unidade) ficam no JSON e no CSV, para quem precisar auditar
de onde veio cada numero.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..config import Config
from ..models import Measurement
from ..runner import Resultado
from . import prefixo_arquivo

_CABECALHO_FILL = PatternFill("solid", fgColor="1F2A44")
_CABECALHO_FONT = Font(color="FFFFFF", bold=True)

# O Excel proibe : \ / ? * [ ] em nome de aba, e limita a 31 caracteres.
_INVALIDO_ABA = re.compile(r"[:\\/?*\[\]]")

COLUNAS = ["Campo", "Valor"]


def gravar(resultado: Resultado, config: Config) -> Path:
    caminho = config.saida_dir / f"{prefixo_arquivo(resultado)}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    usados: set[str] = set()
    for categoria, medicoes in resultado.por_categoria().items():
        ws = wb.create_sheet(_nome_aba(categoria, usados))
        _escrever_aba(ws, medicoes)

    # Um arquivo sem nenhuma aba nao abre no Excel.
    if not wb.sheetnames:
        _escrever_aba(wb.create_sheet("Valores"), [])

    wb.save(caminho)
    return caminho


def _escrever_aba(ws, medicoes: list[Measurement]) -> None:
    ws.append(COLUNAS)
    for celula in ws[1]:
        celula.font = _CABECALHO_FONT
        celula.fill = _CABECALHO_FILL
        celula.alignment = Alignment(vertical="center")

    for m in medicoes:
        # Valor None (query que falhou) vira celula vazia: o campo existe no
        # dashboard, so nao foi possivel ler. Zero seria mentira.
        ws.append([m.nome_valor, m.valor])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUNAS))}{ws.max_row}"
    _ajustar_larguras(ws, medicoes)


def _ajustar_larguras(ws, medicoes: list[Measurement]) -> None:
    maior = len(COLUNAS[0])
    # Amostra as 200 primeiras: varrer tudo nao muda a largura final.
    for m in medicoes[:200]:
        maior = max(maior, len(m.nome_valor))
    ws.column_dimensions["A"].width = min(max(maior + 2, 20), 80)
    ws.column_dimensions["B"].width = 16


def _nome_aba(categoria: str, usados: set[str]) -> str:
    base = _INVALIDO_ABA.sub("-", categoria).strip() or "sem-categoria"
    base = base[:31]
    nome = base
    sufixo = 2
    while nome.lower() in {u.lower() for u in usados}:
        corte = 31 - len(f"_{sufixo}")
        nome = f"{base[:corte]}_{sufixo}"
        sufixo += 1
    usados.add(nome)
    return nome
