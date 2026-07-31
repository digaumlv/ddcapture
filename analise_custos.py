"""Cruza os precos contratados com o uso capturado do dashboard.

    python analise_custos.py

Le precos.sqlite (tabelas de preco) e o JSON mais recente de out/ (captura do
dashboard) e gera analise_custos.xlsx.

DUAS ARMADILHAS que o script trata explicitamente:

1. Os nomes de canal REPETEM em cada etapa do funil (REQUEST, SENT, RETSC,
   RETRC, MREAD...). Somar por nome contaria a mesma mensagem varias vezes.
   Aqui a etapa e extraida da query e vira dimensao propria.

2. As queries usam negacao: '-@log.templateCategory:UTILITY' significa
   WhatsApp SEM a categoria Utility. Um regex ingenuo le isso como se fosse
   'com Utility' e mistura populacoes complementares.
"""

from __future__ import annotations

import json
import re
import sqlite3
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RAIZ = Path(__file__).resolve().parent
BANCO = RAIZ / "precos.sqlite"

# Codigos e nomes dos emissores ficam em config/emissores.txt, fora do
# versionamento - sao dados da carteira, nao do programa.
ARQUIVO_EMISSORES = RAIZ / "config" / "emissores.txt"

# O codigo do emissor no nome do arquivo de captura:
# <dashboard>_1234_20260730-173848.json
_CODIGO_NO_ARQUIVO = re.compile(r"_(\d{4})_\d{8}-\d{6}\.json$")


def nomes_emissores() -> dict[str, str]:
    """Mapa codigo -> nome, so para rotular a planilha.

    Arquivo opcional: sem ele a planilha usa o proprio codigo como rotulo.
    """
    nomes: dict[str, str] = {}
    if not ARQUIVO_EMISSORES.exists():
        return nomes
    for linha in ARQUIVO_EMISSORES.read_text(encoding="utf-8").splitlines():
        linha = linha.strip()
        if not linha or linha.startswith("#"):
            continue
        codigo, _, nome = linha.partition("=")
        codigo = codigo.strip()
        if codigo:
            nomes[codigo] = nome.strip() or codigo
    return nomes

_FILL_CAB = PatternFill("solid", fgColor="1F2A44")
_FONT_CAB = Font(color="FFFFFF", bold=True)
_FONT_ROTULO = Font(bold=True)
_FILL_NOTA = PatternFill("solid", fgColor="FFF4CE")

# O '-' na frente inverte o sentido do filtro - precisa ser capturado.
_CANAL = re.compile(r"(-?)@log\.messageChannel:([A-Za-z]+)")
_CATEG = re.compile(r"(-?)@log\.templateCategory:([A-Z]+)")
_TIPO = re.compile(r"@type:(SIC_\w+)")
_STATUS = re.compile(r"@log\.statusCode:\(?([A-Z]+(?:\s+OR\s+[A-Z]+)*)\)?")

# Status que representa mensagem efetivamente enviada. E a premissa de
# faturamento adotada aqui - trocar esta constante muda o custo calculado.
# A etapa completa e 'RESPONSE SENT'; casamos pelo sufixo para nao depender
# do prefixo do tipo.
STATUS_FATURAVEL = "SENT"


def capturas_por_emissor() -> dict[str, Path]:
    """Mapa codigo -> JSON mais recente daquele emissor.

    So considera arquivos com codigo no nome: uma captura geral misturaria
    os emissores, que e justamente o que se quer evitar aqui.
    """
    encontrados: dict[str, Path] = {}
    for arquivo in sorted((RAIZ / "out").glob("*.json")):
        m = _CODIGO_NO_ARQUIVO.search(arquivo.name)
        if m:
            encontrados[m.group(1)] = arquivo  # ordenado: fica o mais recente
    return encontrados


def _diagnostico_sem_captura() -> str:
    """Explica POR QUE nao ha captura utilizavel, olhando o que existe em out/.

    'Rode a captura' e resposta errada quando a captura rodou: o caso comum e
    o sink json estar desligado, e ai out/ tem xlsx mas nenhum json - e so o
    json carrega a query de cada valor, que a analise precisa.
    """
    pasta = RAIZ / "out"
    if not pasta.exists():
        return (
            f"A pasta {pasta.name}/ nao existe.\n"
            "Rode antes: 3_capturar_emissores.bat"
        )

    arquivos = sorted(p.name for p in pasta.iterdir() if p.is_file())
    if not arquivos:
        return (
            f"A pasta {pasta.name}/ esta vazia.\n"
            "Rode antes: 3_capturar_emissores.bat"
        )

    jsons = [n for n in arquivos if n.lower().endswith(".json")]
    if not jsons:
        outros = ", ".join(sorted({Path(n).suffix or "?" for n in arquivos}))
        return (
            f"Existem {len(arquivos)} arquivo(s) em {pasta.name}/ ({outros}), "
            "mas nenhum .json.\n\n"
            "A analise precisa do sink json: so ele carrega a query de cada\n"
            "valor, que e o que permite saber canal e etapa. O xlsx tem\n"
            "apenas Campo e Valor.\n\n"
            "Ligue em config/settings.yaml (ou settings.local.yaml):\n"
            "    sinks:\n"
            "      json: true\n\n"
            "e rode a captura de novo: 3_capturar_emissores.bat"
        )

    return (
        f"Ha {len(jsons)} json em {pasta.name}/, mas nenhum com codigo de\n"
        f"emissor no nome: {', '.join(jsons[:3])}\n\n"
        "A analise so usa capturas por emissor - uma captura geral misturaria\n"
        "os emissores. Rode: 3_capturar_emissores.bat\n"
        "(o ddcapture.bat sozinho gera captura sem codigo no nome)"
    )


def classificar(query: str) -> tuple[str | None, str]:
    """Devolve (canal, etapa) a partir da query do widget.

    Canal vem com a categoria de template quando houver, e a negacao e
    preservada: 'WhatsApp (exceto UTILITY)' e populacao diferente de
    'WhatsApp/UTILITY'.
    """
    m_canal = _CANAL.search(query)
    if not m_canal:
        return None, ""

    negado, canal = m_canal.group(1), m_canal.group(2)
    if negado:
        canal = f"(exceto {canal})"

    m_categ = _CATEG.search(query)
    if m_categ:
        neg_c, categoria = m_categ.group(1), m_categ.group(2)
        canal += f" (exceto {categoria})" if neg_c else f"/{categoria}"

    m_tipo = _TIPO.search(query)
    m_status = _STATUS.search(query)
    etapa = m_tipo.group(1).replace("SIC_", "") if m_tipo else "?"
    if m_status:
        etapa = f"{etapa} {m_status.group(1)}"
    return canal, etapa


def total_eventos(medicoes: list[dict]) -> tuple[float, str]:
    """Volume de eventos do periodo, para comparar com a franquia da faixa.

    Usa o campo de requisicoes totais: e o topo do funil, antes de qualquer
    quebra por canal ou etapa, entao nao corre risco de dupla contagem.
    """
    candidatos = ("total requisitadas", "1. requisitadas")
    for alvo in candidatos:
        for m in medicoes:
            nome = (m.get("nome_valor") or "").strip().lower()
            if nome == alvo and isinstance(m.get("valor"), (int, float)):
                return float(m["valor"]), m["nome_valor"]
    return 0.0, "(nao encontrado)"


# O widget que lista os canais em uso no periodo. E um group_by: canal sem
# nenhum evento simplesmente nao aparece como grupo.
_WIDGET_CANAIS = re.compile(r"quantidade\s+por\s+canais", re.I)


def _normalizar_canal(texto: str) -> str:
    """'WhatsApp', 'Whatsapp', 'whats app' -> 'whatsapp'."""
    return re.sub(r"[^a-z0-9]", "", unicodedata.normalize("NFKD", texto)
                  .encode("ascii", "ignore").decode().lower())


def canais_com_uso(medicoes: list[dict]) -> dict[str, float]:
    """Canais que aparecem em 'Quantidade Por Canais' com volume > 0.

    E o que decide se o piso fixo do canal e devido: canal sem uso no periodo
    nao gera cobranca. Somamos as linhas do mesmo canal antes de comparar com
    zero, porque o widget quebra o canal em mais de uma linha.
    """
    somas: dict[str, float] = defaultdict(float)
    rotulos: dict[str, str] = {}

    for m in medicoes:
        if not _WIDGET_CANAIS.search(m.get("widget_titulo") or ""):
            continue
        canal, _ = classificar(m.get("query") or "")
        if not canal:
            # O canal tambem vem no nome: 'Quantidade Por Canais - WhatsApp'.
            # Quando a quebra e por group_by, o sufixo vem como facet cru
            # ('@log.messageChannel:WhatsApp') - so o valor apos o ':' serve.
            partes = (m.get("nome_valor") or "").rsplit(" - ", 1)
            canal = partes[1] if len(partes) == 2 else ""
            canal = canal.rsplit(":", 1)[-1]
        canal = canal.split("/")[0].strip()
        if not canal or canal.startswith("("):
            continue
        chave = _normalizar_canal(canal)
        if not chave:
            continue
        rotulos.setdefault(chave, canal)
        valor = m.get("valor")
        if isinstance(valor, (int, float)):
            somas[chave] += valor

    return {rotulos[k]: v for k, v in somas.items() if v > 0}


def _cabecalho(ws, colunas: list[str]) -> None:
    ws.append(colunas)
    for c in ws[ws.max_row]:
        c.font = _FONT_CAB
        c.fill = _FILL_CAB
        c.alignment = Alignment(vertical="center", wrap_text=True)


def _larguras(ws, larguras: list[int]) -> None:
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def aba_fixo(
    wb: Workbook,
    conexao: sqlite3.Connection,
    ativos: dict[str, float],
) -> float:
    """Tudo que e cobrado por valor fixo, independente de volume.

    O piso de um canal so e devido se o canal TEVE USO no periodo. Canal sem
    nenhum evento nao gera cobranca - por isso `ativos`, vindo do widget de
    quantidade por canais.
    """
    ws = wb.create_sheet("Tarifa Fixa")

    ws.append(["Itens de tarifa FIXA - cobrados so quando o canal teve uso"])
    ws["A1"].font = _FONT_ROTULO
    ws.append([])

    _cabecalho(ws, [
        "Origem", "Item", "Volume no periodo", "Cobrado?", "Tarifa (R$)", "Valor (R$)",
    ])

    ativos_norm = {_normalizar_canal(c): v for c, v in ativos.items()}

    total = 0.0
    for canal, tarifa_ini, tarifa_fim in conexao.execute(
        "SELECT canal, tarifa_inicial, tarifa_final FROM piso_canais"
        " WHERE lower(tipo_tarifa)='fixa' ORDER BY tarifa_final DESC"
    ):
        volume = ativos_norm.get(_normalizar_canal(canal))
        if volume:
            total += tarifa_fim or 0
            ws.append(["Piso por canal", canal, volume, "SIM", tarifa_fim, tarifa_fim])
        else:
            ws.append([
                "Piso por canal", canal, 0,
                "nao - sem uso no periodo", tarifa_fim, None,
            ])

    for f_ini, f_fim, t_fin, t_nfin in conexao.execute(
        "SELECT faixa_inicial, faixa_final,"
        " tarifa_eventos_financeiros, tarifa_eventos_nao_financeiros"
        " FROM faixas_preco WHERE lower(tipo_tarifa)='fixa'"
    ):
        faixa = f"ate {f_fim:,.0f} eventos".replace(",", ".")

        # Dentro da franquia, o fixo devido e o de eventos NAO financeiros.
        ws.append([
            "Faixa SIC - ev. nao financeiros", faixa, None, "SIM", t_nfin, t_nfin,
        ])
        total += t_nfin or 0

        # NAO entra no total: a cobranca de eventos financeiros aparece apenas
        # no excedente, por multiplicacao - ver a aba 'Custo variavel'.
        ws.append([
            "Faixa SIC - ev. financeiros", faixa, None,
            "nao - so acima da franquia", t_fin, None,
        ])

    ws.append([])
    ws.append(["", "TOTAL FIXO", "", "", "", total])
    for c in ws[ws.max_row]:
        c.font = _FONT_ROTULO

    ws.append([])
    ws.append([
        "O piso de um canal so entra no total quando o canal teve uso no "
        "periodo. A coluna 'Volume no periodo' vem do widget 'Quantidade Por "
        "Canais' - canal que nao aparece la nao gera cobranca."
    ])
    ws.cell(row=ws.max_row, column=1).fill = _FILL_NOTA
    ws.append([
        "ATENCAO: este valor e do CONTRATO, nao do emissor. Ele aparece igual "
        "na planilha de cada emissor, para referencia - somar as duas planilhas "
        "contaria o piso duas vezes. So o custo variavel e por emissor."
    ])
    ws.cell(row=ws.max_row, column=1).fill = _FILL_NOTA

    _larguras(ws, [32, 30, 18, 26, 16, 16])
    return total


def aba_uso(wb: Workbook, medicoes: list[dict]) -> dict[str, float]:
    """Volume por canal x etapa do funil. A etapa e o que evita duplicar."""
    ws = wb.create_sheet("Uso por canal")

    somas: dict[tuple[str, str], float] = defaultdict(float)
    for m in medicoes:
        canal, etapa = classificar(m.get("query") or "")
        if not canal:
            continue
        valor = m.get("valor")
        if isinstance(valor, (int, float)):
            somas[(canal, etapa)] += valor

    canais = sorted({c for c, _ in somas})
    etapas = sorted({e for _, e in somas})

    ws.append(["Volume por canal e etapa do funil (01/07 a 31/07)"])
    ws["A1"].font = _FONT_ROTULO
    ws.append([])

    _cabecalho(ws, ["Canal"] + etapas + ["TOTAL"])
    for canal in canais:
        linha = [canal] + [somas.get((canal, e), 0) for e in etapas]
        linha.append(sum(linha[1:]))
        ws.append(linha)

    ws.append([])
    ws.append([
        "Nota: o mesmo canal aparece em varias etapas do funil. Somar a linha "
        "TOTAL contaria a mesma mensagem mais de uma vez - use a coluna da "
        "etapa que interessa. 'exceto X' marca filtro negado na query."
    ])
    ws.cell(row=ws.max_row, column=1).fill = _FILL_NOTA

    _larguras(ws, [30] + [14] * len(etapas) + [14])

    # A etapa vem como 'RESPONSE SENT'; casar pelo sufixo evita depender do
    # prefixo do tipo, que ja mudou uma vez.
    return {
        canal: sum(
            v for (c, etapa), v in somas.items()
            if c == canal and etapa.endswith(STATUS_FATURAVEL)
        )
        for canal in canais
    }


def custo_por_faixas(
    volume_total: float, faixas: list[tuple[int | None, int, float, float]]
) -> tuple[float, float, list[tuple[str, float, float, float, float, float]]]:
    """Cobra o volume progressivamente, faixa a faixa.

    Recebe o volume TOTAL, nao o excedente: os limites das faixas ja sao
    absolutos (250.001 a 500.000...), e a primeira faixa variavel comeca
    justo onde a franquia fixa termina. Passar o excedente aqui desloca todas
    as faixas e zera a cobranca.

    Cada faixa cobra so a fatia que cai dentro dela - nao a tarifa da ultima
    faixa sobre o total. E o que explica as tarifas cairem conforme o volume
    sobe.
    """
    detalhe: list[tuple[str, float, float, float, float, float]] = []
    total_fin = total_nfin = 0.0

    for inicio, fim, tarifa_fin, tarifa_nfin in faixas:
        piso = (inicio or 1) - 1
        fatia = max(0.0, min(volume_total, fim) - piso)
        if fatia <= 0:
            continue
        custo_fin = fatia * (tarifa_fin or 0)
        custo_nfin = fatia * (tarifa_nfin or 0)
        total_fin += custo_fin
        total_nfin += custo_nfin
        rotulo = f"{piso + 1:,.0f} a {fim:,.0f}".replace(",", ".")
        detalhe.append(
            (rotulo, fatia, tarifa_fin or 0, custo_fin, tarifa_nfin or 0, custo_nfin)
        )

    return total_fin, total_nfin, detalhe


def aba_custo(
    wb: Workbook,
    conexao: sqlite3.Connection,
    enviados: dict[str, float],
    eventos_sic: float,
) -> float:
    """Quantidade do canal x tarifa da faixa em que o volume caiu.

    'Tarifa Inicial' e 'Tarifa Final' nao sao desconto negociado: sao a tarifa
    ANTES e DEPOIS do limite em 'Valor final' (1.000.000). Ao ultrapassar o
    limite, a tarifa menor passa a valer para a quantidade INTEIRA, nao so
    para o excedente - por isso a coluna 'Tarifa aplicada'.
    """
    ws = wb.create_sheet("Custo variavel")

    ws.append([f"Custo por canal - status considerado: {STATUS_FATURAVEL}"])
    ws["A1"].font = _FONT_ROTULO
    ws.append([])

    _cabecalho(ws, [
        "Canal (captura)", "Canal (tabela de preco)", "Quantidade",
        "Limite (Valor final)", "Faixa", "Tarifa aplicada (R$)", "Custo (R$)",
    ])

    precos = {
        canal.lower(): (canal, valor_final, tarifa_ini, tarifa_fim)
        for canal, valor_final, tarifa_ini, tarifa_fim in conexao.execute(
            "SELECT canal, valor_final, tarifa_inicial, tarifa_final FROM piso_canais"
            " WHERE lower(tipo_tarifa) LIKE 'vari%'"
        )
    }

    def casar(canal_captura: str):
        alvo = canal_captura.lower()
        if "/utility" in alvo:
            chave = "whatsapp utility"
        elif "/marketing" in alvo:
            chave = "whatsapp marketing"
        elif "/authentication" in alvo:
            chave = "whatsapp autenticacao"
        elif alvo.startswith("sms"):
            chave = "sms envio"
        elif alvo.startswith("email"):
            chave = "email"
        elif alvo.startswith("voz"):
            chave = "voz"
        elif "whatsapp" in alvo:
            chave = "whatsapp freeform"
        else:
            return None, None, None, None
        return precos.get(chave) or (None, None, None, None)

    total = 0.0
    for canal, volume in sorted(enviados.items(), key=lambda kv: -kv[1]):
        nome_preco, limite, tarifa_ini, tarifa_fim = casar(canal)
        if limite is None:
            ws.append([canal, "(sem correspondencia)", volume, None, None, None, None])
            continue

        # Passar do limite troca a tarifa da quantidade TODA, nao so do
        # excedente - e um desconto por volume, nao uma faixa progressiva.
        acima = volume > limite
        tarifa = tarifa_fim if acima else tarifa_ini
        faixa = "acima do limite" if acima else "ate o limite"

        custo = volume * (tarifa or 0)
        total += custo
        ws.append([canal, nome_preco, volume, limite, faixa, tarifa, custo])

    ws.append([])
    ws.append(["", "", "", "", "", "TOTAL CANAIS", total])
    for c in ws[ws.max_row]:
        c.font = _FONT_ROTULO

    # --- Eventos SIC: faixa fixa ate 250.000, progressiva acima disso -------
    faixas = list(
        conexao.execute(
            "SELECT faixa_inicial, faixa_final,"
            " tarifa_eventos_financeiros, tarifa_eventos_nao_financeiros"
            " FROM faixas_preco WHERE lower(tipo_tarifa) LIKE 'vari%'"
            " ORDER BY faixa_final"
        )
    )
    franquia_sic = conexao.execute(
        "SELECT MAX(faixa_final) FROM faixas_preco WHERE lower(tipo_tarifa)='fixa'"
    ).fetchone()[0] or 0

    excedente_sic = max(0.0, eventos_sic - franquia_sic)
    # Volume TOTAL, nao o excedente: as faixas ja comecam onde a franquia acaba.
    custo_fin, custo_nfin, detalhe = custo_por_faixas(eventos_sic, faixas)

    ws.append([])
    ws.append(["Eventos SIC - faixa (so o que passa da franquia e cobrado)"])
    ws.cell(row=ws.max_row, column=1).font = _FONT_ROTULO
    _cabecalho(ws, [
        "Faixa", "", "Eventos no periodo", "Franquia (fixa)", "Excedente",
        "Tarifa financeiro", "Custo financeiro", "Tarifa nao financeiro",
        "Custo nao financeiro",
    ])
    ws.append([
        "(total do periodo)", "", eventos_sic, franquia_sic, excedente_sic,
        None, None, None, None,
    ])

    for rotulo, fatia, tarifa, custo, tarifa_n, custo_n in detalhe:
        ws.append([rotulo, "", None, None, fatia, tarifa, custo, tarifa_n, custo_n])

    ws.append(["", "", "", "", "", "", custo_fin, "", custo_nfin])
    ws.cell(row=ws.max_row, column=1).value = "TOTAL EVENTOS SIC"
    for c in ws[ws.max_row]:
        c.font = _FONT_ROTULO

    ws.append([])
    geral = total + custo_fin + custo_nfin
    ws.append(["", "", "", "", "", "", "", "TOTAL VARIAVEL GERAL", geral])
    for c in ws[ws.max_row]:
        c.font = _FONT_ROTULO

    ws.append([])
    ws.append([
        "Canais: a quantidade INTEIRA e multiplicada pela tarifa da faixa em "
        "que o volume caiu. Ate 1.000.000 vale a Tarifa Inicial; acima disso a "
        "Tarifa Final passa a valer para tudo, nao so para o excedente."
    ])
    ws.cell(row=ws.max_row, column=1).fill = _FILL_NOTA
    ws.append([
        "PREMISSAS: status " + STATUS_FATURAVEL + " como faturavel; coluna "
        "'Tarifa Final'; faixas SIC cobradas de forma progressiva (cada faixa "
        "cobra so a fatia que cai nela). O casamento canal/preco e por nome."
    ])
    ws.cell(row=ws.max_row, column=1).fill = _FILL_NOTA

    _larguras(ws, [30, 26, 18, 20, 14, 18, 18, 20, 20])
    return total + custo_fin + custo_nfin


def aba_detalhe(wb: Workbook, medicoes: list[dict]) -> None:
    """Todo campo com canal, para auditar de onde saiu cada numero."""
    ws = wb.create_sheet("Detalhe")
    _cabecalho(ws, ["Canal", "Etapa", "Campo", "Valor", "Query"])

    linhas = []
    for m in medicoes:
        canal, etapa = classificar(m.get("query") or "")
        if canal:
            linhas.append([canal, etapa, m["nome_valor"], m.get("valor"), m.get("query")])

    for linha in sorted(linhas, key=lambda l: (l[0], l[1])):
        ws.append(linha)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _larguras(ws, [26, 20, 34, 12, 100])


def analisar(codigo: str, arquivo: Path, conexao: sqlite3.Connection) -> dict:
    """Gera a planilha de custo de UM emissor."""
    dados = json.loads(arquivo.read_text(encoding="utf-8"))
    medicoes = dados["medicoes"]
    nome = nomes_emissores().get(codigo, codigo)

    wb = Workbook()
    wb.remove(wb.active)

    eventos, origem = total_eventos(medicoes)
    ativos = canais_com_uso(medicoes)
    total_fixo = aba_fixo(wb, conexao, ativos)
    enviados = aba_uso(wb, medicoes)
    total_variavel = aba_custo(wb, conexao, enviados, eventos)
    aba_detalhe(wb, medicoes)

    # Aba de capa: sem ela, duas planilhas iguais no layout ficam faceis de
    # confundir - e o codigo do emissor so apareceria no nome do arquivo.
    capa = wb.create_sheet("Emissor", 0)
    janela = dados["janela"]
    for rotulo, valor in (
        ("Emissor", f"{nome} ({codigo})"),
        ("Filtro aplicado", f"@org:({codigo} OR {codigo.lstrip('0')})"),
        ("Dashboard", f"{dados['dashboard']['titulo']} ({dados['dashboard']['id']})"),
        ("Janela", f"{janela['from'][:10]} a {janela['to'][:10]}"),
        ("Eventos no periodo", eventos),
        ("Origem dos eventos", origem),
        (
            "Canais com uso",
            ", ".join(f"{c} ({v:,.0f})" for c, v in sorted(ativos.items()))
            or "(nenhum)",
        ),
        ("", ""),
        ("Custo VARIAVEL do emissor", total_variavel),
        ("Tarifa fixa do CONTRATO", total_fixo),
    ):
        capa.append([rotulo, valor])
        capa.cell(row=capa.max_row, column=1).font = _FONT_ROTULO
    capa.append([])
    capa.append([
        "A tarifa fixa e do contrato, nao deste emissor: aparece igual nas duas "
        "planilhas. Somar as duas contaria o piso em dobro. So o custo variavel "
        "e atribuivel ao emissor."
    ])
    capa.cell(row=capa.max_row, column=1).fill = _FILL_NOTA
    _larguras(capa, [30, 52])

    saida = RAIZ / f"analise_custos_{codigo}.xlsx"
    try:
        wb.save(saida)
    except PermissionError:
        # O Excel trava o arquivo aberto. Um traceback aqui esconderia a
        # causa, que tem solucao trivial.
        raise SystemExit(
            f"\nNao foi possivel gravar {saida.name}: o arquivo esta aberto.\n"
            "Feche-o no Excel e rode de novo."
        ) from None

    return {
        "codigo": codigo,
        "nome": nome,
        "arquivo": saida.name,
        "eventos": eventos,
        "fixo": total_fixo,
        "variavel": total_variavel,
    }


def main() -> int:
    if not BANCO.exists():
        raise SystemExit("precos.sqlite nao encontrado. Rode: python importar_precos.py")

    capturas = capturas_por_emissor()
    if not capturas:
        raise SystemExit(_diagnostico_sem_captura())

    conexao = sqlite3.connect(BANCO)
    try:
        resumos = [
            analisar(codigo, arquivo, conexao)
            for codigo, arquivo in sorted(capturas.items())
        ]
    finally:
        conexao.close()

    print(f"{len(resumos)} planilha(s) geradas:\n")
    for r in resumos:
        print(f"  {r['arquivo']}")
        print(f"     eventos        : {r['eventos']:>12,.0f}")
        print(f"     custo variavel : R$ {r['variavel']:>10,.2f}")
        print()

    print(f"  Soma do variavel : R$ {sum(r['variavel'] for r in resumos):,.2f}")
    print(f"  Fixo do contrato : R$ {resumos[0]['fixo']:,.2f}  (nao somar por emissor)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
