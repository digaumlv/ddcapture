"""Saida: planilha por emissor e resumo no console."""

from __future__ import annotations

import logging
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .configuracao import RAIZ
from .precificador import ResultadoEmissor

log = logging.getLogger(__name__)

_FILL_CAB = PatternFill("solid", fgColor="1F2A44")
_FONT_CAB = Font(color="FFFFFF", bold=True)
_FONT_ROTULO = Font(bold=True)
_FILL_ERRO = PatternFill("solid", fgColor="FDE7E9")
_FILL_NOTA = PatternFill("solid", fgColor="FFF4CE")

COLUNAS = [
    "Tipo", "Servico", "Quantidade", "Regra aplicada", "Valor (R$)", "Erro",
]


def gerar(resultado: ResultadoEmissor) -> Path:
    """Uma planilha por emissor. Devolve o caminho gravado."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Precificacao"

    for rotulo, valor in (
        ("Emissor", f"{resultado.emissor} ({resultado.codigo})"),
        ("Janela", resultado.janela),
        ("", ""),
        ("Total SIC", resultado.por_tipo("SIC")),
        ("Total FIXO", resultado.por_tipo("FIXO")),
        ("Total VARIAVEL", resultado.por_tipo("VARIAVEL")),
        ("TOTAL", resultado.total),
    ):
        ws.append([rotulo, _celula(valor)])
        ws.cell(row=ws.max_row, column=1).font = _FONT_ROTULO

    ws.append([])
    linha_cab = ws.max_row + 1
    ws.append(COLUNAS)
    for celula in ws[linha_cab]:
        celula.font = _FONT_CAB
        celula.fill = _FILL_CAB
        celula.alignment = Alignment(vertical="center", wrap_text=True)

    for item in resultado.itens:
        ws.append([
            item.tipo,
            item.servico,
            _celula(item.quantidade),
            item.regra,
            _celula(item.valor),
            item.erro or "",
        ])
        if not item.ok:
            for celula in ws[ws.max_row]:
                celula.fill = _FILL_ERRO

    ws.append([])
    ws.append([
        "Valor vazio com erro preenchido significa que o item nao pode ser"
        " calculado. Nao vira zero: zero e um valor financeiro legitimo e"
        " mascararia o problema no total."
    ])
    ws.cell(row=ws.max_row, column=1).fill = _FILL_NOTA

    ws.freeze_panes = f"A{linha_cab + 1}"
    for i, largura in enumerate((12, 38, 14, 46, 16, 40), start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura

    caminho = RAIZ / f"precificacao_{resultado.codigo}.xlsx"
    try:
        wb.save(caminho)
    except PermissionError:
        # O Excel trava o arquivo aberto; um traceback esconderia uma causa
        # de solucao trivial.
        raise SystemExit(
            f"\nNao foi possivel gravar {caminho.name}: o arquivo esta aberto."
            "\nFeche-o no Excel e rode de novo."
        ) from None

    log.info("planilha gravada: %s", caminho.name)
    return caminho


def _celula(valor):
    if valor is None:
        return ""
    if isinstance(valor, Decimal):
        return float(valor)
    return valor


def resumo(resultados: list[ResultadoEmissor], arquivos: list[Path]) -> None:
    """Resumo consolidado no console."""
    print()
    print("=" * 74)
    print("  RESUMO DA PRECIFICACAO")
    print("=" * 74)

    cab = f"  {'Emissor':<24}{'SIC':>12}{'FIXO':>12}{'VARIAVEL':>12}{'TOTAL':>14}"
    print(cab)
    print("  " + "-" * (len(cab) - 2))

    geral = Decimal("0")
    erros = 0
    for r in resultados:
        print(
            f"  {r.emissor[:22] + ' (' + r.codigo + ')':<24}"
            f"{r.por_tipo('SIC'):>12,.2f}{r.por_tipo('FIXO'):>12,.2f}"
            f"{r.por_tipo('VARIAVEL'):>12,.2f}{r.total:>14,.2f}"
        )
        geral += r.total
        erros += len(r.erros)

    print("  " + "-" * (len(cab) - 2))
    print(f"  {'TOTAL GERAL':<24}{'':>36}{geral:>14,.2f}")

    if erros:
        print(f"\n  {erros} item(ns) nao calculado(s):")
        for r in resultados:
            for item in r.erros:
                print(f"    [{r.codigo}] {item.servico}: {item.erro}")

    print("\n  Planilhas:")
    for caminho in arquivos:
        print(f"    {caminho.name}")
